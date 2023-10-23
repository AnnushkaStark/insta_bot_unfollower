
#Этот файл для загрузки exel таблицы любого объема ы базу данных


import sqlite3 as sql
from openpyxl import load_workbook
from database_controller import add_accounts
file  = 'selected_by_categories.xlsx'

def add_questions_from_excel(file):
    ''' Загрузка пользователей в базу данных из exel'''
    try:    # Попытка открыть файл
        workbook = load_workbook(filename = file) # Загружаем Excel-файл
        sheet = workbook['DataSheet']  # Указываем название страницы (можно ли переделать на индекс?)'
        if sheet["B1"].value == "Instagram" and sheet["C1"].value == "Link" and sheet["D1"].value =='Instagram ID' and sheet["E1"].value == 'Name':      # Проверка на соответствие загружаемого файла шаблону
            for row in sheet.iter_rows(min_row=2, values_only=True):
                add_accounts(row[1], row[2],row[3],row[4])
            return 'Аккаунты добавлены'
            
        else:
            return 'Файл не соответствует формату'
            
    except ValueError:
          return'error'
print(add_questions_from_excel(file))